import csv
import pandas as pd
from openpyxl import load_workbook
import numpy
import matplotlib.pyplot as plt
from sqlalchemy import create_engine
# from infocards.archive import Archive

#Example 1 Raw Data conversion

#with open('rawdata_example1.csv') as csvfile:
#	reader = csv.DictReader(csvfile)
#	list1 = []
#	for row in reader:
#		if(row['kW'] != ''):  #ignores Nan types at the end of csv list (in excel format)
#			list1.append(row['kW'])


#Assign filename to file
file1 = 'rawdata_example1.csv'

#Load spreadsheet
#x1 = pd.read_csv(file1, header=None, skiprows=1, dtype={'Date-time':object, 'kW':int}, low_memory=False)
x1 = pd.read_csv(file1, header=None, skiprows=1, low_memory=False)

list1 = []
for x in x1[1]:
	if isinstance(x, str):
		if x == ' ':
			continue
		else:
			list1.append(int(x))
			#print(x)
			#continue		
	elif x is None or pd.isnull(x) :
		continue		

#-------------------------------------------------------------------

#Example 2 raw data conversion

#Assign filename to file
file2 = 'rawdata_example2.xlsx'

#Load spreadsheet
x2 = pd.ExcelFile(file2)

df = x2.parse(x2.sheet_names[0])

values = df['Site kWh'].values
list2 = []
count = 0
for value in values:
	count = count +1
	#convert kwH to kw
	#then append to vector/list four times for each 15min interval
	# as long as value exists
	if(pd.notnull(value)):
		quarter_interval= value/0.25
		list2.append(quarter_interval) 
		list2.append(quarter_interval) 
		list2.append(quarter_interval) 
		list2.append(quarter_interval) 
		#print(quarter_interval)

#print (count)

#we should delete the last 4 from list2 because they contain misinforming data from the total consumption in the kWh column
list2 = list2[:-4]

#-------------------------------------------------------------------

#Example 3 raw data conversion
file3 = 'rawdata_example3.xlsx'

#Load spreadsheet
#x3 = pd.ExcelFile(file3)

#Load excel workbook
wb = load_workbook(file3)

sheet = wb.sheetnames

# Get sheet by name
sheet = wb[sheet[0]]

#df = x3.parse(x3.sheet_names[0])

list3= []
#iterate through each row contianing kWh data
for i in range(3 , sheet.max_row, 3):
	for j in range(5, sheet.max_column , 3):
		# get the mean of the next 3 energy values
		a1 = sheet.cell(row=i, column = j).value
		#check for missing data
		if a1 is None or isinstance(a1,str):
			a1 = 0

		a2 = sheet.cell(row=i, column = j+1).value
		if a2 is None or isinstance(a2,str):
			a2 = 0

		a3 = sheet.cell(row=i, column = j+2).value
		if a3 is None or isinstance(a3,str):
			a3 = 0

		a = [a1, a2, a3]
		mean = numpy.mean(a)
		
		# convert energy back to power before adding to list
		kW = mean/0.25
		list3.append(kW)


#----------------------------------------------------------------------
#Creating Visualizations
fig=plt.figure()
sp1 = fig.add_subplot(3,1,1)
sp2 = fig.add_subplot(3,1,2)
sp3 = fig.add_subplot(3,1,3)


# initialize vector/listoflist
v1 = []
v2 = []
v3 = []

#declare vectors independently
time_frame1 = []
for i in range (0,len(list1)):
	time_frame1.append(0.25*i)
	v1.append([time_frame1[i], list1[i]])

time_frame2 = []
for i in range (0,len(list2)):
	time_frame2.append(0.25*i)
	v1.append([time_frame2[i], list2[i]])

time_frame3 = []
for i in range (0,len(list3)):
	time_frame3.append(0.25*i)
	v3.append([time_frame3[i], list3[i]])

#vector1 = numpy.zeros(2,dtype={'names':('tf', 'power' ),'formats':('float', 'float')})

sp1.scatter(time_frame1,list1, marker='o', s=0.5)
sp2.scatter(time_frame2,list2, marker='o', s=0.5)
sp3.scatter(time_frame3,list3, marker='o', s=0.5)

#plt.title('Power consumption relative to days in 15 min intervals- Example1')
plt.xlabel('Time in 15-min intervals')
plt.ylabel('Power Consumption')
plt.show()


#-----------------------------------------------------------------------
#Creating a database engine
 
engine = create_engine('sqlite:///EDF_Client_data.sqlite')

#arc = Archive(
#    db_name='my_db',
#    db_type='mysql',
#    user='mysql_user',
#    password='mysql_password',
#    host='mysql_host',
#    port=1234
#)

#title = 'EDF-RE client archiving'
#description'Archiving power consumption'
#content = vector1 + vector2 + vector3
#tags = vectors
#arc.new_card(title, description, content, tags)

